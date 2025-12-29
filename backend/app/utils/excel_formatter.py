"""
Utilidades para formatear archivos Excel con estilo profesional.
"""
import io
from typing import Optional, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
    NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# Colores corporativos
HEADER_BG_COLOR = "1E3A5F"  # Azul oscuro
HEADER_FONT_COLOR = "FFFFFF"  # Blanco
ALT_ROW_COLOR = "F5F7FA"  # Gris muy claro para filas alternadas
BORDER_COLOR = "D1D5DB"  # Gris para bordes


def create_formatted_excel(
    df: pd.DataFrame,
    sheet_name: str = "Datos",
    title: Optional[str] = None,
    column_config: Optional[Dict[str, Dict[str, Any]]] = None
) -> io.BytesIO:
    """
    Crea un archivo Excel formateado profesionalmente.

    Args:
        df: DataFrame con los datos a exportar
        sheet_name: Nombre de la hoja
        title: Título opcional (no se usa, mantenido por compatibilidad)
        column_config: Configuración opcional de columnas {nombre: {width: int, format: str}}

    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Estilos
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_row_fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    cell_alignment = Alignment(vertical="center", wrap_text=False)

    start_row = 1

    # Escribir headers
    for col_idx, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[start_row].height = 25

    # Escribir datos
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Filas alternadas
            if (row_idx - start_row) % 2 == 0:
                cell.fill = alt_row_fill

    # Ajustar ancho de columnas
    for col_idx, column_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_idx)

        # Calcular ancho basado en contenido
        max_length = len(str(column_name))

        for row in ws.iter_rows(min_row=start_row + 1, max_row=min(start_row + 100, ws.max_row),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except:
                    pass

        # Limitar el ancho máximo y mínimo
        adjusted_width = min(max(max_length + 2, 10), 50)

        # Usar configuración personalizada si existe
        if column_config and column_name in column_config:
            config = column_config[column_name]
            if "width" in config:
                adjusted_width = config["width"]

        ws.column_dimensions[column_letter].width = adjusted_width

    # Congelar la fila de headers
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Agregar filtros automáticos
    if len(df) > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{ws.max_row}"

    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def get_column_config_nncc() -> Dict[str, Dict[str, Any]]:
    """Configuración de columnas para informe NNCC."""
    return {
        "cliente": {"width": 15},
        "nombre_cliente": {"width": 25},
        "direccion": {"width": 35},
        "comuna": {"width": 15},
        "zona": {"width": 12},
        "inspector": {"width": 20},
        "fecha_inspeccion": {"width": 14},
        "estado_efectividad": {"width": 18},
        "observaciones": {"width": 40},
    }


def get_column_config_lecturas() -> Dict[str, Dict[str, Any]]:
    """Configuración de columnas para informe de Lecturas."""
    return {
        "numero_cliente": {"width": 15},
        "nombre": {"width": 25},
        "direccion": {"width": 35},
        "sector": {"width": 12},
        "inspector": {"width": 20},
        "fecha": {"width": 14},
        "hallazgo": {"width": 20},
        "origen": {"width": 15},
        "observaciones": {"width": 40},
    }


def get_column_config_teleco() -> Dict[str, Dict[str, Any]]:
    """Configuración de columnas para informe de Telecomunicaciones."""
    return {
        "empresa": {"width": 20},
        "comuna": {"width": 15},
        "direccion": {"width": 35},
        "resultado": {"width": 18},
        "tiene_plano": {"width": 14},
        "fecha": {"width": 14},
        "observaciones": {"width": 40},
    }


def get_column_config_calidad() -> Dict[str, Dict[str, Any]]:
    """Configuración de columnas para informe de Control de Pérdidas."""
    return {
        "numero_cliente": {"width": 15},
        "nombre": {"width": 25},
        "direccion": {"width": 35},
        "comuna": {"width": 15},
        "tipo_sistema": {"width": 18},
        "tipo_resultado": {"width": 18},
        "contratista": {"width": 20},
        "inspector": {"width": 20},
        "fecha": {"width": 14},
        "observaciones": {"width": 40},
    }


def get_column_config_corte() -> Dict[str, Dict[str, Any]]:
    """Configuración de columnas para informe de Corte y Reposición."""
    return {
        "suministro": {"width": 15},
        "nombre": {"width": 25},
        "direccion": {"width": 35},
        "comuna": {"width": 15},
        "zona": {"width": 12},
        "centro_operativo": {"width": 18},
        "inspector": {"width": 20},
        "situacion_encontrada": {"width": 22},
        "motivo_multa": {"width": 20},
        "fecha": {"width": 14},
        "observaciones": {"width": 40},
    }
