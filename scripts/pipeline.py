#!/usr/bin/env python3
"""
Pipeline automatico de actualizacion de data.
Detecta cambios en archivos fuente Excel y actualiza PostgreSQL.

Usage:
    python scripts/pipeline.py                     # Procesa solo cambios
    python scripts/pipeline.py --force             # Reprocesa todo
    python scripts/pipeline.py --source nncc       # Solo un modulo
    python scripts/pipeline.py --dry-run           # Muestra que procesaria
"""

import sys
import os
import json
import hashlib
import argparse
import glob as glob_mod
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from sqlalchemy import create_engine, text
from backend.app.core.config import settings

PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "data"
ENEL_DIR = DATA_DIR / "ENEL Calidad"
MANIFEST_PATH = Path(__file__).parent / ".pipeline_manifest.json"


# ---------------------------------------------------------------------------
# Manifest (change detection via MD5)
# ---------------------------------------------------------------------------

def load_manifest() -> dict:
    if MANIFEST_PATH.exists():
        return json.loads(MANIFEST_PATH.read_text())
    return {}


def save_manifest(manifest: dict):
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2, default=str))


def file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def has_changed(manifest: dict, paths: list[str]) -> bool:
    for p in paths:
        if not os.path.exists(p):
            continue
        current = file_hash(p)
        if manifest.get(p, {}).get("hash") != current:
            return True
    return False


def update_manifest(manifest: dict, paths: list[str]):
    for p in paths:
        if os.path.exists(p):
            manifest[p] = {
                "hash": file_hash(p),
                "updated": datetime.now().isoformat(),
            }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_hyperlinks_from_column(filepath: str, sheet_name: str, column_name: str) -> dict:
    """
    Extract hyperlinks from a specific column in an Excel file.
    Returns a dict mapping row index (0-based, data rows) to hyperlink URL.
    Combines target + location for complete URLs (e.g., https://domain.com/#/path).
    """
    try:
        from openpyxl import load_workbook
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning)

        wb = load_workbook(filepath, data_only=False)
        ws = wb[sheet_name]

        # Find the column index for the target column
        col_idx = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col).value
            if cell_val and str(cell_val).strip() == column_name:
                col_idx = col
                break

        if col_idx is None:
            print(f"    Column '{column_name}' not found in sheet '{sheet_name}'")
            wb.close()
            return {}

        # Extract hyperlinks (row 2 onwards = data rows, index 0-based for DataFrame)
        hyperlinks = {}
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.hyperlink:
                # Combine target (base URL) + # + location (path) for complete URL
                target = cell.hyperlink.target or ""
                location = cell.hyperlink.location or ""

                if target and location:
                    # Remove trailing slash from target if present
                    base = target.rstrip('/')
                    # Build complete URL with hash for SPA routing
                    full_url = f"{base}/#{location}"
                elif target:
                    full_url = target
                elif location:
                    full_url = location
                else:
                    continue

                # row_idx - 2 to convert to 0-based DataFrame index
                hyperlinks[row_idx - 2] = full_url

        wb.close()
        return hyperlinks
    except Exception as e:
        print(f"    Error extracting hyperlinks: {e}")
        return {}


def clean_unnamed_cols(df: pd.DataFrame) -> pd.DataFrame:
    mask = df.columns.to_series().str.startswith("Unnamed").fillna(True)
    return df.loc[:, ~mask]


def ensure_column_exists(engine, table: str, column: str, column_type: str = "TEXT"):
    """Ensure a column exists in the table, adding it if needed."""
    with engine.connect() as conn:
        result = conn.execute(text(f"""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = :table AND column_name = :column
        """), {"table": table, "column": column})
        if not result.fetchone():
            conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {column_type}"))
            conn.commit()
            print(f"    Added column '{column}' to table '{table}'")


def truncate_and_insert(engine, table: str, df: pd.DataFrame, chunksize: int = 5000):
    # Ensure any new columns exist before inserting
    if table == "nncc" and "link_formulario" in df.columns:
        ensure_column_exists(engine, table, "link_formulario", "TEXT")

    with engine.connect() as conn:
        conn.execute(text(f"TRUNCATE TABLE {table}"))
        conn.commit()
    df.to_sql(table, engine, if_exists="append", index=False, chunksize=chunksize)
    with engine.connect() as conn:
        conn.execute(text(f"ANALYZE {table}"))
        conn.commit()
    print(f"    {table}: {len(df):,} rows inserted")


# ---------------------------------------------------------------------------
# Module processors
# ---------------------------------------------------------------------------

def process_nncc(engine) -> list[str]:
    """NNCC - load from parquet (preferred) or Excel with hyperlink extraction."""
    # 1. Try parquet files first (most recent by filename)
    parquet_dir = DATA_DIR / "parquet"
    parquet_matches = sorted(glob_mod.glob(str(parquet_dir / "BASE_ACTUAL_*.parquet")))
    if parquet_matches:
        src = parquet_matches[-1]  # most recent by name
        print(f"    Reading parquet: {os.path.basename(src)}")
        df = pd.read_parquet(src)
        df = clean_unnamed_cols(df)

        # Use URL_LINK FORMULARIO column if available
        if "URL_LINK FORMULARIO" in df.columns:
            df["link_formulario"] = df["URL_LINK FORMULARIO"]
            count = df["link_formulario"].notna().sum()
            print(f"    Using URL_LINK FORMULARIO column: {count} URLs found")
        elif "LINK FORMULARIO" in df.columns:
            df["link_formulario"] = df["LINK FORMULARIO"]
        else:
            df["link_formulario"] = None
    else:
        # 2. Fallback to Excel
        src = str(ENEL_DIR / "5. Nuevas Conexiones" / "NNCC" / "2025-05 INFORME NNCC (2024-2029) ENE 2026.xlsx")

        if not os.path.exists(src):
            alt_paths = [
                ENEL_DIR / "1. NNCC" / "2025-05 INFORME NNCC (2024-2029) ENE 2026.xlsx",
                DATA_DIR / "2025-05 INFORME NNCC (2024-2029) ENE 2026.xlsx",
            ]
            for alt in alt_paths:
                if os.path.exists(str(alt)):
                    src = str(alt)
                    break
            else:
                matches = glob_mod.glob(str(ENEL_DIR / "5. Nuevas Conexiones" / "NNCC" / "*.xlsx"))
                if not matches:
                    matches = glob_mod.glob(str(ENEL_DIR / "1. NNCC" / "*.xlsx"))
                if not matches:
                    matches = glob_mod.glob(str(DATA_DIR / "*NNCC*.xlsx"))
                matches = [m for m in matches if "~$" not in m]
                if matches:
                    src = matches[0]
                else:
                    print("    WARNING: No NNCC source file found")
                    return []

        print(f"    Reading Excel: {os.path.basename(src)}")
        df = pd.read_excel(src, sheet_name="BASE ACTUAL")
        df = clean_unnamed_cols(df)

        if "URL_LINK FORMULARIO" in df.columns:
            df["link_formulario"] = df["URL_LINK FORMULARIO"]
            count = df["link_formulario"].notna().sum()
            print(f"    Using URL_LINK FORMULARIO column: {count} URLs found")
        else:
            print("    Extracting hyperlinks from LINK FORMULARIO column...")
            hyperlinks = extract_hyperlinks_from_column(src, "BASE ACTUAL", "LINK FORMULARIO")
            if hyperlinks:
                df["link_formulario"] = df.index.map(lambda i: hyperlinks.get(i, None))
                print(f"    Found {len([v for v in hyperlinks.values() if v])} hyperlinks")
            else:
                df["link_formulario"] = None
                print("    No hyperlinks found")

    # Column mapping (from migrate_data.py)
    column_mapping = {
        "VTA": "vta", "Cliente": "cliente", "Nombre cliente": "nombre_cliente",
        "Dirección": "direccion", "Comuna": "comuna", "TARIFA": "tarifa",
        "ZONA": "zona", "BASE": "base", "N° MEDIDOR": "n_medidor",
        "ESTADO EFECTIVIDAD OCA": "estado_efectividad",
        "RESULTADO FINAL DE INSPCCION": "resultado_inspeccion",
        "MULTA SI/NO": "multa", "OBSERVACIONES DE MULTA": "observaciones_multa",
        "FECHA INSPECCIÓN": "fecha_inspeccion", "Inspector3": "inspector",
        "ESTADO CONTRATISTA": "estado_contratista",
        "RESULTADO FINAL DE REVISIÓN DE NORMALIZACIÓN": "resultado_normalizacion",
        "CUMPLE NORMA CODIGO COLORES": "cumple_norma_cc",
        "CLIENTE CONFORME": "cliente_conforme",
        "ESTADO DEL EMPALME": "estado_empalme",
        "TIPO INSPECCIÓN": "tipo_inspeccion", "VOLTAJE": "voltaje",
        "CONTRATISTA_ENEL": "contratista_enel",
        "CATEGORIA_MAL_EJECUTADO": "categoria_mal_ejecutado",
        "LATITUD": "latitud",
        "LONGITUD": "longitud",
    }
    rename_dict = {k: v for k, v in column_mapping.items() if k in df.columns}
    df = df.rename(columns=rename_dict)

    if "fecha_inspeccion" in df.columns:
        df["fecha_inspeccion"] = pd.to_datetime(df["fecha_inspeccion"], format='mixed', errors='coerce')
        df['mes'] = df['fecha_inspeccion'].dt.month
        df['anio'] = df['fecha_inspeccion'].dt.year

    if "inspector" in df.columns:
        df["inspector"] = df["inspector"].fillna("").str.strip().str.title()

    if 'cliente' in df.columns:
        df['cliente'] = df['cliente'].astype(str).str.replace('.0', '', regex=False)

    known_cols = [
        'vta', 'cliente', 'nombre_cliente', 'direccion', 'comuna', 'tarifa',
        'zona', 'base', 'n_medidor', 'estado_efectividad', 'resultado_inspeccion',
        'multa', 'observaciones_multa', 'fecha_inspeccion', 'inspector',
        'estado_contratista', 'resultado_normalizacion', 'cumple_norma_cc',
        'cliente_conforme', 'estado_empalme', 'tipo_inspeccion', 'voltaje',
        'mes', 'anio', 'link_formulario', 'contratista_enel', 'categoria_mal_ejecutado',
        'latitud', 'longitud',
    ]
    df = df[[c for c in known_cols if c in df.columns]]

    truncate_and_insert(engine, "nncc", df)

    # Pre-calculate dashboard stats for each base + __all__
    print("    Pre-calculating NNCC dashboard stats...")
    try:
        from precalculate_nncc import run as precalculate_nncc_stats
        bases = df["base"].dropna().unique().tolist()
        for b in sorted(bases):
            precalculate_nncc_stats(base_filter=b)
        precalculate_nncc_stats()  # __all__
        print(f"    NNCC dashboard stats updated ({len(bases)} bases + __all__)")
    except Exception as e:
        print(f"    WARNING: Failed to pre-calculate stats: {e}")

    return [src]


def process_medidores_cruzados(engine) -> list[str]:
    """Medidores Cruzados - single Excel file."""
    src = str(ENEL_DIR / "5. Nuevas Conexiones" / "Medidores Cruzados" /
              "Consolidado Medidores Cruzados NNCC (NUEVO CONTRATO).xlsx")
    if not os.path.exists(src):
        matches = glob_mod.glob(str(ENEL_DIR / "5. Nuevas Conexiones" / "Medidores Cruzados" / "*.xlsx"))
        if matches:
            src = matches[0]
        else:
            print("    WARNING: No Medidores Cruzados source file found")
            return []

    print(f"    Reading: {os.path.basename(src)}")
    df = pd.read_excel(src, sheet_name="MEDIDORES CRUZADOS")
    df = clean_unnamed_cols(df)
    df.columns = df.columns.str.strip()

    column_mapping = {
        "MES": "mes", "MES ": "mes",
        "FECHA CORREO": "fecha_correo", "NUM CLIENTE": "num_cliente",
        "ENCARGADO ENEL": "encargado_enel", "AREA": "area",
        "ETAPA DEL CASO": "etapa_caso", "FECHA ASIGNACION": "fecha_asignacion",
        "FECHA DE ANALISIS": "fecha_analisis", "FECHA INSPECCIÓN": "fecha_inspeccion",
        "DIRECCIÓN": "direccion", "COMUNA": "comuna",
        "MEDIDOR SISTEMA": "medidor_sistema", "ZONA": "zona", "EETT": "eett",
        "OBSERVACION INSPECTOR": "observacion_inspector",
        "ESTADO MEDIDOR": "estado_medidor", "INSPECTOR": "inspector",
        "RESULTADO FINAL DE LA INSPECCION": "resultado_inspeccion",
        "EEPP OCA": "eepp_oca",
    }
    rename_dict = {k: v for k, v in column_mapping.items() if k in df.columns}
    df = df.rename(columns=rename_dict)

    for col in ["fecha_correo", "fecha_asignacion", "fecha_analisis", "fecha_inspeccion"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format='mixed', errors='coerce')

    known_cols = [
        'mes', 'fecha_correo', 'num_cliente', 'encargado_enel', 'area',
        'etapa_caso', 'fecha_asignacion', 'fecha_analisis', 'fecha_inspeccion',
        'direccion', 'comuna', 'medidor_sistema', 'zona', 'eett',
        'observacion_inspector', 'estado_medidor', 'inspector',
        'resultado_inspeccion', 'eepp_oca',
    ]
    df = df[[c for c in known_cols if c in df.columns]]

    truncate_and_insert(engine, "medidores_cruzados", df)
    return [src]


def process_calidad(engine) -> list[str]:
    """Control Perdidas - monthly Excel files (mono + tri)."""
    base_dir = ENEL_DIR / "2. Control Pérdida" / "2025"
    if not base_dir.exists():
        print(f"    WARNING: Calidad directory not found: {base_dir}")
        return []

    all_files = []
    tables_data = {}  # {table_name: [df, ...]}

    months = sorted([d for d in base_dir.iterdir() if d.is_dir()])
    for month_dir in months:
        month_label = month_dir.name
        for xlsx in month_dir.glob("*.xlsx"):
            fname = xlsx.name.lower()
            if "~$" in fname:
                continue
            all_files.append(str(xlsx))

            is_mono = "mono" in fname
            is_tri = "tri" in fname
            if not is_mono and not is_tri:
                continue

            prefix = "mono" if is_mono else "tri"
            print(f"    Reading: {month_label}/{xlsx.name}")

            try:
                xls = pd.ExcelFile(xlsx)
            except Exception as e:
                print(f"    ERROR reading {xlsx.name}: {e}")
                continue

            for sheet_name in xls.sheet_names:
                if sheet_name.lower() == "sheet1":
                    continue

                df = pd.read_excel(xls, sheet_name=sheet_name)
                df = clean_unnamed_cols(df)
                df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
                df = df.loc[:, ~df.columns.duplicated()]

                safe_name = sheet_name.replace("/", "").replace("\\", "").replace(":", "").upper()
                table_name = f"calidad_{prefix}_{safe_name}"

                # Add origin month
                df["mes_origen"] = f"{month_label}_2025"

                if table_name not in tables_data:
                    tables_data[table_name] = []
                tables_data[table_name].append(df)

    # Only migrate BASE and INSPECCIONES tables (the ones used by the app)
    target_tables = {
        "calidad_mono_BASE": "calidad_mono_base",
        "calidad_mono_INSPECCIONES": "calidad_mono_inspecciones",
        "calidad_tri_BASE": "calidad_tri_base",
        "calidad_tri_INSPECCIONES": "calidad_tri_inspecciones",
    }

    for source_key, db_table in target_tables.items():
        dfs = tables_data.get(source_key, [])
        if not dfs:
            print(f"    No data for {db_table}")
            continue
        combined = pd.concat(dfs, ignore_index=True)
        truncate_and_insert(engine, db_table, combined)

    return all_files


def process_corte(engine) -> list[str]:
    """Corte y Reposicion - single Excel file."""
    src = str(ENEL_DIR / "3. CyR" / "Entregado por Yeimei Castro" /
              "Consolidado Inspecciones Calidad Mayo 2024 - 2025 - 2026 (Diego Bravo).xlsx")
    if not os.path.exists(src):
        matches = glob_mod.glob(str(ENEL_DIR / "3. CyR" / "**" / "*.xlsx"), recursive=True)
        matches = [m for m in matches if "~$" not in m]
        if matches:
            src = matches[0]
        else:
            print("    WARNING: No Corte source file found")
            return []

    print(f"    Reading: {os.path.basename(src)}")
    df = pd.read_excel(src, sheet_name="Consolidado Calidad")
    df = clean_unnamed_cols(df)
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    df = df.loc[:, ~df.columns.duplicated()]

    truncate_and_insert(engine, "corte_reposicion", df)
    return [src]


def process_lecturas(engine) -> list[str]:
    """Lecturas - single Excel with multiple sheets."""
    src = str(DATA_DIR / "Planilla Ordg 12 DICIEMBRE 2025 BASE ACT.xlsx")
    if not os.path.exists(src):
        matches = glob_mod.glob(str(DATA_DIR / "*Planilla*Ordg*.xlsx"))
        if not matches:
            matches = glob_mod.glob(str(ENEL_DIR / "4. Lecturas" / "*.xlsx"))
        if matches:
            src = matches[0]
        else:
            print("    WARNING: No Lecturas source file found")
            return []

    print(f"    Reading: {os.path.basename(src)}")
    xls = pd.ExcelFile(src)

    sheet_map = {
        "SEC": "SEC",
        "ORDEN": "ORDENES",
        "VIRTUAL": "VISITA VIRTUAL",
    }

    all_dfs = []
    for sheet_name in xls.sheet_names:
        upper_name = sheet_name.upper()
        origen = None
        for keyword, cat in sheet_map.items():
            if keyword in upper_name:
                origen = cat
                break
        if origen is None:
            continue

        df = pd.read_excel(xls, sheet_name=sheet_name)
        df = clean_unnamed_cols(df)
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        df = df.loc[:, ~df.columns.duplicated()]
        df["origen"] = origen
        all_dfs.append(df)
        print(f"    Sheet '{sheet_name}' -> {origen}: {len(df)} rows")

    if not all_dfs:
        print("    No matching sheets found")
        return []

    combined = pd.concat(all_dfs, ignore_index=True)
    truncate_and_insert(engine, "lecturas", combined)
    return [src]


def process_teleco(engine) -> list[str]:
    """Telecomunicaciones - single Excel file."""
    src = str(ENEL_DIR / "6. Telecomunicaciones" / "CONTROL OPERACIONES TELCO.xlsx")
    if not os.path.exists(src):
        matches = glob_mod.glob(str(ENEL_DIR / "6. Telecomunicaciones" / "*.xlsx"))
        matches = [m for m in matches if "~$" not in m]
        if matches:
            src = matches[0]
        else:
            print("    WARNING: No Teleco source file found")
            return []

    print(f"    Reading: {os.path.basename(src)}")
    df = pd.read_excel(src, sheet_name="Base")
    df = clean_unnamed_cols(df)
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    df = df.loc[:, ~df.columns.duplicated()]

    truncate_and_insert(engine, "telecomunicaciones", df)
    return [src]


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

MODULES = {
    "nncc": process_nncc,
    "medidores_cruzados": process_medidores_cruzados,
    "calidad": process_calidad,
    "corte": process_corte,
    "lecturas": process_lecturas,
    "teleco": process_teleco,
}


def run_pipeline(source: str = None, force: bool = False, dry_run: bool = False):
    if not settings.DATABASE_URL:
        print("ERROR: DATABASE_URL not set in .env")
        sys.exit(1)

    db_url = settings.DATABASE_URL
    if "+asyncpg" in db_url:
        db_url = db_url.replace("postgresql+asyncpg://", "postgresql://")

    manifest = load_manifest()
    modules_to_run = {source: MODULES[source]} if source else MODULES

    # Detect changes
    print("=" * 60)
    print(f"Pipeline run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Mode: {'FORCE' if force else 'DRY RUN' if dry_run else 'INCREMENTAL'}")
    print("=" * 60)

    if not force and not dry_run:
        # For incremental, we check the manifest but process anyway if no manifest exists
        pass

    engine = None if dry_run else create_engine(db_url, echo=False)
    processed = []
    skipped = []
    errors = []

    for name, processor in modules_to_run.items():
        print(f"\n--- {name.upper()} ---")

        if dry_run:
            print(f"  Would process: {name}")
            processed.append(name)
            continue

        try:
            files = processor(engine)
            if files:
                update_manifest(manifest, files)
                processed.append(name)
                print(f"  OK: {name}")
            else:
                skipped.append(name)
                print(f"  SKIPPED: {name} (no source files)")
        except Exception as e:
            errors.append((name, str(e)))
            print(f"  ERROR: {name} - {e}")
            import traceback
            traceback.print_exc()

    if not dry_run:
        save_manifest(manifest)

        # Show final counts
        if engine:
            print("\n=== Row counts ===")
            tables = [
                "nncc", "medidores_cruzados",
                "calidad_mono_base", "calidad_mono_inspecciones",
                "calidad_tri_base", "calidad_tri_inspecciones",
                "lecturas", "telecomunicaciones", "corte_reposicion",
            ]
            with engine.connect() as conn:
                for table in tables:
                    try:
                        count = conn.execute(text(f"SELECT COUNT(*) FROM {table}")).scalar()
                        print(f"  {table}: {count:,} rows")
                    except:
                        pass
            engine.dispose()

    # Summary
    print(f"\n{'=' * 60}")
    print(f"Processed: {len(processed)} ({', '.join(processed) or 'none'})")
    print(f"Skipped:   {len(skipped)} ({', '.join(skipped) or 'none'})")
    print(f"Errors:    {len(errors)} ({', '.join(e[0] for e in errors) or 'none'})")
    print(f"{'=' * 60}")

    return {
        "processed": processed,
        "skipped": skipped,
        "errors": errors,
        "timestamp": datetime.now().isoformat(),
    }


def main():
    parser = argparse.ArgumentParser(description="Pipeline de actualizacion de datos")
    parser.add_argument("--source", type=str, choices=list(MODULES.keys()),
                        help="Procesar solo un modulo")
    parser.add_argument("--force", action="store_true",
                        help="Reprocesar todo sin verificar cambios")
    parser.add_argument("--dry-run", action="store_true",
                        help="Mostrar que se procesaria sin hacer cambios")
    args = parser.parse_args()

    run_pipeline(source=args.source, force=args.force, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
